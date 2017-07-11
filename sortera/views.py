# -*- coding: utf-8 -*-

'''
Created on 7 apr. 2017

@author: perhk
'''

from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from .forms import UploadFileForm

import csv, io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Font

files_to_download = {}


def mk_one_file(rows, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = filename[18:23]  # !!!!!!!!!!!!!!!!!!!
    bold_font=Font(bold=True)
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    columns = ["Tidpunkt event", "SystemId", "AnvändarID", "Händelsetyp", "ObjektId", "Relation", "Objektinfo", "Reserv 1", "Reserv 2", "Anstalldes arbetsort", "Anstalldes Af-kontor", "Anstalldes narmsta chef", "Anstalldes MO", "Anstalldes Mo/Avd -chef"]
    header = ["Tidpunkt", "System", "Signatur", "Händelsetyp", "Objekt", "Relation", "Info 1", "Info 2", "Info 3", "Anställdes arbetsort", "Af-kontor", "Närmsta chef", "Anställdes MO/Avd", "MO/Avd-chef"]

    colsizes = [20,10,8,12,14,55,50,45,24,20,29,13,33,13]
    for col in range(len(header)):
        ws.cell(row=1,column=col+1).value = header[col]
        ws.cell(row=1,column=col+1).font = bold_font
        ws.cell(row=1,column=col+1).fill = yellow_fill
        ws.column_dimensions[chr(65+col)].width = colsizes[col]
    r = 2
    for row in rows:
        c = 1
        for h in columns:
            if h == "Tidpunkt event":
                ws.cell(row=r,column=c).value = datetime.strptime(row[h], "%b %d %Y %H:%M:%S") 
#                     ws.cell(row=r,column=c).number_format = r"YYYY-MM-DD HH:MM:SS"
                ws.cell(row=r,column=c).number_format = r"yyyy/mm/dd hh:mm:ss;@"
            else:
                if h == "Händelsetyp":
                    row[h] = {"read":"Läsa","create":"Skapa","update":"Uppdatera","delete":"ta bort","search":"Söka"}[row[h].lower()]
                if h in row:
                    ws.cell(row=r,column=c).value = row[h]
            c += 1
        r += 1
        files_to_download[filename+".xlsx"] = io.BytesIO(save_virtual_workbook(wb))

def rd_data(f):
    rows = []
    reader = csv.DictReader(io.StringIO(f.read().decode('utf-8')))
    for row in reader:
        if rows and rows[0]['AnvändarID']!=row["AnvändarID"]:
            d = datetime.strptime(rows[0]["Tidpunkt event"], "%b %d %Y %H:%M:%S")
            s = d.strftime("%Y-%m-%d")
            fname = 'Anhörigslagningar {} {}'.format(rows[0]['AnvändarID'], s)
            mk_one_file(rows, fname)
            rows = []
        rows.append(row)
    if rows:
        d = datetime.strptime(rows[0]["Tidpunkt event"], "%b %d %Y %H:%M:%S")
        s = d.strftime("%Y-%m-%d")
        fname = 'Anhörigslagningar {} {}'.format(rows[0]['AnvändarID'], s)
        mk_one_file(rows, fname)
        rows = []

def ladda_upp(request):
    files_to_download.clear()
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            rd_data(request.FILES['file'])
            return HttpResponseRedirect('/ladda-ner-lista')
    else:
        form = UploadFileForm()
    return render(request, 'ladda-upp.html', {'form': form})


def ladda_ner_lista(request):
    flist = []
    for f in files_to_download:
        flist.append(f)
    return render(request, 'ladda-ner-lista.html', {"fil_lista":flist})

def ladda_ner_fil(request, filnamn):
#     print(filnamn)
    response = HttpResponse(files_to_download[filnamn], content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = "attachment; filename="+filnamn
    return response



